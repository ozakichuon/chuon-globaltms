#!/usr/bin/env python3
"""
中温従業員情報の2Excelから全データを抽出してJSONへ。
- ■技能実習生等一覧.xlsx（管理表・履歴・寮管理・資格管理表・入社予定等）
- 資格点数一覧ピポットテーブル26.3.5～.xlsx（Sheet1生データ）

Output: src/lib/data/*.json
"""
import json
import os
import re
import uuid
from datetime import datetime, date
import openpyxl

BASE = '/Users/yuyu24/2ndBrain/chuon-tms'
XLSX_MAIN = f'{BASE}/従業員情報/■技能実習生等一覧.xlsx'
XLSX_CERT = f'{BASE}/従業員情報/資格点数一覧ピポットテーブル26.3.5～.xlsx'
OUT_DIR = f'{BASE}/src/lib/data'
os.makedirs(OUT_DIR, exist_ok=True)


def d(v):
    """日付変換: datetime/date → 'YYYY-MM-DD', 数値(Excelシリアル)も処理"""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        # Excelシリアル値の可能性（1900/1/1起点、ただし1900年閏日バグあり）
        if 20000 < v < 60000:
            try:
                base = datetime(1899, 12, 30)
                from datetime import timedelta
                return (base + timedelta(days=int(v))).strftime('%Y-%m-%d')
            except Exception:
                return None
    # 文字列で '2025/09', '2025/9月' のような半端なもの
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', s)
        if m:
            return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        m = re.match(r'^(\d{4})/(\d{1,2})$', s)
        if m:
            return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-01"
        # '9999.9.9' のような取得日不明
        if s.startswith('9999'):
            return None
    return None


def s(v):
    """文字列化・None保持"""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip() or None


def n(v):
    """数値化"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', ''))
    except Exception:
        return None


# UUID生成をコード安定にするため固定シード
def make_id(prefix, key):
    ns = uuid.NAMESPACE_URL
    return str(uuid.uuid5(ns, f'chuon:{prefix}:{key}'))


# =============================================================
# メイン抽出
# =============================================================
wb = openpyxl.load_workbook(XLSX_MAIN, data_only=True)

# -------------------------
# 1. 管理表 → employees + visa_records
# -------------------------
ws = wb['管理表']
employees = []
visa_records = []
support_agencies = set()

for r in range(2, ws.max_row + 1):
    name = s(ws.cell(r, 8).value)
    if not name:
        continue
    code = s(ws.cell(r, 5).value)
    if not code:
        # 社員コード無しの行は、生まれたばかりの子ども等の特殊ケース。スキップ
        continue
    nationality_jp = s(ws.cell(r, 4).value)
    nat_map = {'インドネシア': 'ID', 'ベトナム': 'VN', 'フィリピン': 'PH', '日本': 'JP'}
    nationality = nat_map.get(nationality_jp, 'OTHER')

    visa_type_jp = s(ws.cell(r, 17).value)
    visa_map = {
        '技能実習1号': 'technical_intern_1',
        '技能実習2号': 'technical_intern_2',
        '技能実習3号': 'technical_intern_3',
        '特定技能１号': 'specified_skill_1',
        '特定技能1号': 'specified_skill_1',
        '特定技能２号': 'specified_skill_2',
        '特定技能2号': 'specified_skill_2',
        '特定活動': 'other',
    }
    visa_status = visa_map.get(visa_type_jp, 'other')

    retired_mark = s(ws.cell(r, 34).value)
    retired_at = d(ws.cell(r, 35).value)
    is_retired = retired_mark == '退職' or bool(retired_at)

    gender_jp = s(ws.cell(r, 10).value)
    gender_map = {'男': 'male', '女': 'female'}
    gender = gender_map.get(gender_jp)

    jlpt_raw = s(ws.cell(r, 28).value)
    jlpt = None
    if jlpt_raw:
        m = re.search(r'N[1-5]', jlpt_raw)
        if m:
            jlpt = m.group(0)

    agency = s(ws.cell(r, 14).value)
    if agency:
        support_agencies.add(agency)

    emp_id = make_id('employee', code)
    hired_at = d(ws.cell(r, 15).value)
    emp = {
        'id': emp_id,
        'employee_code': code,
        'display_name': name,
        'furigana': s(ws.cell(r, 7).value),
        'nickname': s(ws.cell(r, 9).value),
        'nationality': nationality,
        'nationality_jp': nationality_jp,
        'gender': gender,
        'birth_date': d(ws.cell(r, 11).value),
        'marital_status': s(ws.cell(r, 13).value),
        'support_agency': agency,
        'hired_at': hired_at,
        'expected_return': s(ws.cell(r, 16).value),  # '2025/09'形式混在のためテキスト保持
        'visa_type_jp': visa_type_jp,
        'visa_status': visa_status,
        'address_jp': s(ws.cell(r, 18).value),
        'room_type': s(ws.cell(r, 19).value),
        'rent_burden': n(ws.cell(r, 20).value),
        'residence_card_no': s(ws.cell(r, 23).value),
        'residence_card_expires_at': d(ws.cell(r, 22).value),
        'employment_insurance_no': s(ws.cell(r, 24).value),
        'jlpt_level': jlpt,
        'department_head': s(ws.cell(r, 37).value),
        'temporary_retirement_date': d(ws.cell(r, 30).value),
        'temporary_return_from': d(ws.cell(r, 31).value),
        'temporary_return_to': d(ws.cell(r, 32).value),
        'rejoined': s(ws.cell(r, 33).value),
        'retired': is_retired,
        'retired_mark': retired_mark,
        'retired_at': retired_at,
        'return_cost': n(ws.cell(r, 36).value),
        'ss1_insurance': s(ws.cell(r, 38).value),
        'ss1_entry_date': d(ws.cell(r, 39).value),
        'ss1_insurance_period': s(ws.cell(r, 40).value),
        'salary_account': s(ws.cell(r, 41).value),
        'note': s(ws.cell(r, 42).value),
        'photo_path': s(ws.cell(r, 6).value),
        'workplace': s(ws.cell(r, 3).value),
        'section': s(ws.cell(r, 1).value),
        'no': n(ws.cell(r, 2).value),
    }
    employees.append(emp)

    if visa_status != 'other' or emp['residence_card_no']:
        visa_records.append({
            'id': make_id('visa', code),
            'employee_id': emp_id,
            'visa_status': visa_status,
            'residence_card_no': emp['residence_card_no'],
            'expires_at': emp['residence_card_expires_at'],
            'is_current': not is_retired,
        })

with open(f'{OUT_DIR}/employees.json', 'w', encoding='utf-8') as f:
    json.dump(employees, f, ensure_ascii=False, indent=2)
with open(f'{OUT_DIR}/visa_records.json', 'w', encoding='utf-8') as f:
    json.dump(visa_records, f, ensure_ascii=False, indent=2)

print(f'employees: {len(employees)} (active {sum(1 for e in employees if not e["retired"])}, retired {sum(1 for e in employees if e["retired"])})')
print(f'visa_records: {len(visa_records)}')
print(f'support_agencies: {len(support_agencies)}: {sorted(support_agencies)}')

# -------------------------
# 2. 登録支援機関マスタ
# -------------------------
agencies = [
    {'id': make_id('agency', a), 'name': a} for a in sorted(support_agencies) if a
]
with open(f'{OUT_DIR}/support_agencies.json', 'w', encoding='utf-8') as f:
    json.dump(agencies, f, ensure_ascii=False, indent=2)

# -------------------------
# 3. 資格マスタ + 資格取得履歴（外部Excel）
# -------------------------
wb2 = openpyxl.load_workbook(XLSX_CERT, data_only=True)
ws2 = wb2['Sheet1']
certifications = {}  # name -> {id, name}
emp_certs = []
# 氏名→employee_idの対応（同姓同名対策で 氏名+勤務地 で突き合わせ）
# ただし、管理表には日本人はほぼ居らず、資格点数Excelは日本人中心の可能性
# 紐付けは「display_nameの空白除去で完全一致」を試みる
name_to_ids = {}
for e in employees:
    key = (e['display_name'] or '').replace(' ', '').replace('　', '')
    name_to_ids.setdefault(key, []).append(e['id'])

unmatched_names = set()
for r in range(2, ws2.max_row + 1):
    name = s(ws2.cell(r, 1).value)
    qname = s(ws2.cell(r, 2).value)
    acq_date = d(ws2.cell(r, 3).value)
    points = n(ws2.cell(r, 4).value)
    prepost = s(ws2.cell(r, 5).value)  # 入社前/入社後
    loc = s(ws2.cell(r, 6).value)
    if not name or not qname:
        continue
    # 資格マスタ
    if qname not in certifications:
        certifications[qname] = {
            'id': make_id('certification', qname),
            'name': qname,
        }
    cert_id = certifications[qname]['id']
    # 紐付け
    key = (name or '').replace(' ', '').replace('　', '')
    emp_ids = name_to_ids.get(key, [])
    emp_id = emp_ids[0] if len(emp_ids) == 1 else None
    if emp_id is None:
        unmatched_names.add(name)
    emp_certs.append({
        'id': make_id('emp_cert', f'{name}:{qname}:{acq_date}'),
        'employee_id': emp_id,  # null許容（未紐付け日本人等）
        'employee_name': name,
        'certification_id': cert_id,
        'certification_name': qname,
        'acquired_at': acq_date,
        'points': points,
        'acquired_before_hire': prepost == '入社前',
        'workplace': loc,
    })

with open(f'{OUT_DIR}/certifications.json', 'w', encoding='utf-8') as f:
    json.dump(list(certifications.values()), f, ensure_ascii=False, indent=2)
with open(f'{OUT_DIR}/employee_certifications.json', 'w', encoding='utf-8') as f:
    json.dump(emp_certs, f, ensure_ascii=False, indent=2)

print(f'certifications: {len(certifications)}')
print(f'employee_certifications: {len(emp_certs)} (matched {sum(1 for c in emp_certs if c["employee_id"])}, unmatched {sum(1 for c in emp_certs if not c["employee_id"])})')
print(f'  unmatched unique names: {len(unmatched_names)}')

# -------------------------
# 4. 生活サポート履歴（履歴シート）
# -------------------------
ws3 = wb['履歴']
support_tickets = []
for r in range(2, ws3.max_row + 1):
    seq = n(ws3.cell(r, 1).value)
    dte = d(ws3.cell(r, 2).value)
    tm = ws3.cell(r, 3).value  # time型
    recorder = s(ws3.cell(r, 4).value)
    target_code = s(ws3.cell(r, 5).value)
    place = s(ws3.cell(r, 6).value)
    kind = s(ws3.cell(r, 7).value)
    done = s(ws3.cell(r, 8).value)
    note_req = s(ws3.cell(r, 11).value)
    responses = []
    # 対応1-5 のペア（日付+対応者+記事）を抽出
    pairs = [(12, 14, 17), (18, 20, 23), (24, 26, 29), (30, 32, 35), (36, 38, 41)]
    for di, ri, ni in pairs:
        rd = d(ws3.cell(r, di).value)
        rr = s(ws3.cell(r, ri).value)
        rn = s(ws3.cell(r, ni).value)
        if rd or rr or rn:
            responses.append({
                'date': rd,
                'responder': rr,
                'note': rn,
            })
    if not target_code and not note_req:
        continue
    # 紐付け: target_codeで
    emp_id = None
    if target_code:
        emp_id = make_id('employee', target_code) if any(e['employee_code'] == target_code for e in employees) else None
    support_tickets.append({
        'id': make_id('ticket', f'{seq}'),
        'seq': int(seq) if seq else None,
        'date': dte,
        'time': str(tm) if tm else None,
        'recorder': recorder,
        'target_employee_code': target_code,
        'target_employee_id': emp_id,
        'place': place,
        'kind': kind,
        'status': done,
        'request_note': note_req,
        'responses': responses,
        'overall_note': s(ws3.cell(r, 42).value),
    })

with open(f'{OUT_DIR}/support_tickets.json', 'w', encoding='utf-8') as f:
    json.dump(support_tickets, f, ensure_ascii=False, indent=2)

print(f'support_tickets: {len(support_tickets)}')

# -------------------------
# 5. 寮マスタ + 部屋 + 入居（10シート統合）
# -------------------------
dorm_sheets = [
    ('小栗部屋(アトム)', 'アトム', '和泉北1-20-28'),
    ('津吉部屋(コーポ津吉)', 'コーポ津吉', '松山市津吉町535-1'),
    ('津吉部屋(ビレッジ)', 'ビレッジハウス松山上野', '松山市上野町774'),
    ('中野①', '中野①', '松山市中野町442-1'),
    ('中野②', '中野②', '松山市中野町457-1'),
    ('中野③', '中野③', '松山市中野町368-6'),
    ('牛渕170-3', '牛渕1号館', '東温市牛渕170-3'),
    ('牛渕123-1', '牛渕2号館', '東温市牛渕123-1'),
    ('西条寮', '西条寮', '西条市丹原町田野上方455-1'),
]
dormitories = []
rooms = []
assignments = []
for sh_name, dorm_name, address in dorm_sheets:
    if sh_name not in wb.sheetnames:
        continue
    ws = wb[sh_name]
    dorm_id = make_id('dormitory', dorm_name)
    dormitories.append({
        'id': dorm_id,
        'name': dorm_name,
        'address': address,
        'sheet_source': sh_name,
    })
    # ヘッダ位置はシートによって違う（ほぼ行3〜4）
    # 列位置は: A=室, B=No, C=勤務地, D=国籍, E=フリガナ, F=性別, G=生年月日, H=年齢,
    # I=登録支援機関, J=入社日, K=実習区分, L=家賃負担, N=部屋番号
    header_row = None
    for r in range(1, 6):
        if s(ws.cell(r, 1).value) == '室' or s(ws.cell(r, 1).value) == '室' or s(ws.cell(r, 1).value) == '間取':
            header_row = r
            break
    if header_row is None:
        header_row = 3
    for r in range(header_row + 1, min(ws.max_row + 1, header_row + 50)):
        name = s(ws.cell(r, 5).value) or s(ws.cell(r, 8).value)
        room_no = s(ws.cell(r, 14).value)
        if not name and not room_no:
            continue
        room_id = make_id('room', f'{dorm_name}:{room_no}:{r}')
        rooms.append({
            'id': room_id,
            'dormitory_id': dorm_id,
            'dormitory_name': dorm_name,
            'room_no': room_no,
            'room_type_label': s(ws.cell(r, 1).value),
        })
        if name:
            # name → employeeとの突合（フリガナで）
            # 管理表のfuriganaとマッチ
            emp_id = None
            for e in employees:
                ek = (e.get('furigana') or '').replace(' ', '').replace('　', '')
                nk = (name or '').replace(' ', '').replace('　', '')
                if ek and nk and (ek == nk or ek in nk or nk in ek):
                    emp_id = e['id']
                    break
            assignments.append({
                'id': make_id('assignment', f'{dorm_name}:{room_no}:{name}:{r}'),
                'dormitory_id': dorm_id,
                'dormitory_name': dorm_name,
                'room_id': room_id,
                'room_no': room_no,
                'resident_name': name,
                'employee_id': emp_id,
                'rent_burden': n(ws.cell(r, 12).value),
                'workplace': s(ws.cell(r, 3).value),
                'nationality': s(ws.cell(r, 4).value),
                'visa_type': s(ws.cell(r, 11).value),
            })

with open(f'{OUT_DIR}/dormitories.json', 'w', encoding='utf-8') as f:
    json.dump(dormitories, f, ensure_ascii=False, indent=2)
with open(f'{OUT_DIR}/rooms.json', 'w', encoding='utf-8') as f:
    json.dump(rooms, f, ensure_ascii=False, indent=2)
with open(f'{OUT_DIR}/dormitory_assignments.json', 'w', encoding='utf-8') as f:
    json.dump(assignments, f, ensure_ascii=False, indent=2)

print(f'dormitories: {len(dormitories)}, rooms: {len(rooms)}, assignments: {len(assignments)}')

# -------------------------
# 6. 日本人指導員の資格管理（資格管理表）
# -------------------------
ws4 = wb['資格管理表']
manager_certs = []
for r in range(6, ws4.max_row + 1):
    name = s(ws4.cell(r, 2).value)
    if not name:
        continue
    # 3種類の資格
    for cert_name, acq_col, exp_col in [
        ('技能実習責任者', 3, 4),
        ('技能実習指導員', 5, 6),
        ('生活指導員', 7, 8),
    ]:
        acq = s(ws4.cell(r, acq_col).value)
        exp = s(ws4.cell(r, exp_col).value)
        if acq or exp:
            manager_certs.append({
                'id': make_id('mgr_cert', f'{name}:{cert_name}'),
                'name': name,
                'certification_name': cert_name,
                'acquired_raw': acq,
                'expires_raw': exp,
            })
with open(f'{OUT_DIR}/manager_certifications.json', 'w', encoding='utf-8') as f:
    json.dump(manager_certs, f, ensure_ascii=False, indent=2)
print(f'manager_certifications: {len(manager_certs)}')

# -------------------------
# 7. 入社予定者（採用パイプライン）
# -------------------------
ws5 = wb['入社予定']
pipeline = []
for r in range(3, min(ws5.max_row + 1, 100)):
    name = s(ws5.cell(r, 8).value) or s(ws5.cell(r, 7).value)
    if not name:
        continue
    pipeline.append({
        'id': make_id('pipeline', f'{r}:{name}'),
        'workplace': s(ws5.cell(r, 3).value),
        'nationality': s(ws5.cell(r, 4).value),
        'furigana': s(ws5.cell(r, 7).value),
        'display_name': name,
        'nickname': s(ws5.cell(r, 9).value),
        'gender': s(ws5.cell(r, 10).value),
        'birth_date': d(ws5.cell(r, 11).value),
        'age': n(ws5.cell(r, 12).value),
        'marital_status': s(ws5.cell(r, 13).value),
        'support_agency': s(ws5.cell(r, 14).value),
        'expected_hire_date_raw': s(ws5.cell(r, 15).value),
        'expected_return_raw': s(ws5.cell(r, 16).value),
        'visa_type_jp': s(ws5.cell(r, 17).value),
        'address_jp': s(ws5.cell(r, 18).value),
        'note': s(ws5.cell(r, 38).value),
    })
with open(f'{OUT_DIR}/recruitment_pipeline.json', 'w', encoding='utf-8') as f:
    json.dump(pipeline, f, ensure_ascii=False, indent=2)
print(f'recruitment_pipeline: {len(pipeline)}')

# -------------------------
# 8. 購入明細（寮備品）
# -------------------------
ws6 = wb['購入明細']
purchases = []
for r in range(2, min(ws6.max_row + 1, 500)):
    dt = d(ws6.cell(r, 1).value)
    factory = s(ws6.cell(r, 2).value)
    house = s(ws6.cell(r, 3).value)
    cat = s(ws6.cell(r, 4).value)
    product = s(ws6.cell(r, 5).value)
    if not product:
        continue
    purchases.append({
        'id': make_id('purchase', f'{r}:{product}:{dt}'),
        'purchased_at': dt,
        'factory': factory,
        'house': house,
        'category': cat,
        'product_name': product,
        'maker': s(ws6.cell(r, 6).value),
        'model': s(ws6.cell(r, 7).value),
        'store': s(ws6.cell(r, 8).value),
        'amount': n(ws6.cell(r, 9).value),
        'warranty': s(ws6.cell(r, 10).value),
        'warranty_end': d(ws6.cell(r, 11).value),
    })
with open(f'{OUT_DIR}/purchases.json', 'w', encoding='utf-8') as f:
    json.dump(purchases, f, ensure_ascii=False, indent=2)
print(f'purchases: {len(purchases)}')

print('\n=== 完了 ===')
for fn in sorted(os.listdir(OUT_DIR)):
    fp = f'{OUT_DIR}/{fn}'
    size = os.path.getsize(fp)
    print(f'  {fn}: {size:,} bytes')
